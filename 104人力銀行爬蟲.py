import requests
from bs4 import BeautifulSoup
import openpyxl
import time
import random
res = requests.get('https://www.104.com.tw/jobs/search/?ro=0&kwop=7&keyword=%E5%A4%A7%E6%95%B8%E6%93%9A&expansionType=area%2Cspec%2Ccom%2Cjob%2Cwf%2Cwktm&order=12&asc=0&page=0&mode=s&jobsource=2018indexpoc&langFlag=0&langStatus=0&recommendJob=1&hotJob=1',verify = False)
soup = BeautifulSoup(res.text)
print('DONE')
from datetime import datetime
today = datetime.today().strftime("%Y-%m-%d")
page_url_front = driver.current_url.split('page')[0]
page_url_behind = driver.current_url.split('page')[1].split('&mode')[1]
page = 1
work = []
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = '職缺名稱'
ws['B1'] = '職缺連結'
ws['C1'] = '公司名稱'
ws['D1'] = '工作地區'
ws['E1'] = '薪資待遇'
ws['F1'] = '給薪方式'
ws['G1'] = '薪資下限'
ws['H1'] = '薪資上限'
while(soup.find_all("article",class_ ="b-block--top-bord job-list-item b-clearfix js-job-item") != []):
    print('===============================\n正在讀取第',page,'頁\n===============================')
    for i in soup.find_all("article",class_ ="b-block--top-bord job-list-item b-clearfix js-job-item"):
        a = (i.a.text)#print("工作名稱:" + i.a.text)
        b = ("https:" + i.a["href"])#print("https:" + i.a["href"])
        c = (i.find('ul',class_ = "b-list-inline b-clearfix").a.text.strip())#print(i.find('ul',class_ = "b-list-inline b-clearfix").a.text.strip())#i.find('ul',class_ = "b-list-inline b-clearfix").a['title'].split()[0]
        d = (i.select('ul')[1].li.text)#print(i.select('ul')[1].li.text)#i.find('ul',class_ = "b-list-inline b-clearfix").a['title'].split()[1]
        if i.find('div',class_ = 'job-list-tag b-content').select('span')!=[] and i.find('div',class_ = 'job-list-tag b-content').span.text == "待遇面議":
            e = (i.find('div',class_ = 'job-list-tag b-content').span.text)#print("工作薪水:" + i.find('div',class_ = 'job-list-tag b-content').span.text)
            f = ''
            g = ''
            h = ''
        else:
            e = (i.find('a',class_ = "b-tag--default").text)#print("工作薪水:" + i.find('a',class_ = "b-tag--default").text)
            f = e[:2]
            word = ''
            for char in e:
                if char == '~' or char.isdigit():
                    word += char
            if '~' in word:
                g = word.split('~')[0]
                h = word.split('~')[1]
            else:
                g = word
                h = word
            g = int(g)
            h = int(h)
        ws.append([a,b,c,d,e,f,g,h])
    page+=1
    res = requests.get(page_url_front+'page='+str(page)+'&mode'+page_url_behind,verify = False)
    soup = BeautifulSoup(res.text)
    wb.save("爬蟲_104人力銀行"+today+".xlsx")